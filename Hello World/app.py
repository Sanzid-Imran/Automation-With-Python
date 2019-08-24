#  this code is to process spreadsheets including fixing one of their column values
#   and show a graph or bar chart following the corresponding corrected values.


import  openpyxl as xl
from openpyxl.chart import Reference, BarChart


def process_workbook(filename):
    wb= xl.load_workbook(filename)
    sheet=wb['Sheet1']

    # In this following block you can change your operation logic as per your demand.

    for row in range(2, sheet.max_row+1):
        cell= sheet.cell(row,3)
        corrected_price=cell.value *0.9
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=corrected_price

    # Picking up the values for making the chart

    value =Reference(sheet,
              min_row=2,
              max_row=4,
              min_col=4,
              max_col=4)

    chart =BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

process_workbook('transactions.xlsx')  #calling the method